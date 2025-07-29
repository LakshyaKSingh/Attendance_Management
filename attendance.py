from flask import Flask, render_template, request, jsonify
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime, timedelta

app = Flask(__name__)

MASTER_DATA_PATH = "E:/Codes/Attendance management/Warehouse_Employees_MasterData.xlsx"
ATTENDANCE_LOG_PATH = "E:/Codes/Attendance management/Attendance_Log.xlsx"

def get_date_options():
    today = datetime.today()
    yesterday = today - timedelta(days=1)
    return today.strftime("%Y-%m-%d"), yesterday.strftime("%Y-%m-%d")  # Use ISO format for <input type=date>

@app.route("/")
def index():
    today, yesterday = get_date_options()
    if os.path.exists(ATTENDANCE_LOG_PATH):
        book = load_workbook(ATTENDANCE_LOG_PATH)
        sheet_dates = book.sheetnames
    else:
        sheet_dates = []
    return render_template("index.html", today=today, yesterday=yesterday, sheet_dates=sheet_dates)

@app.route("/get_employee", methods=["POST"])
def get_employee():
    emp_id = request.json.get("emp_id", "").strip()
    df = pd.read_excel(MASTER_DATA_PATH)
    # Normalize Employee ID column as string and strip spaces
    df["Employee ID"] = df["Employee ID"].astype(str).str.strip()
    match = df[df["Employee ID"] == emp_id]
    if not match.empty:
        data = match.iloc[0]
        return jsonify({
            "found": True,
            "name": data["Full Name"],
            "department": data["Department"],
            "position": data["Position"],
            "status": data["Status"]
        })
    else:
        return jsonify({"found": False, "error": "Employee not found"}), 404

@app.route("/submit", methods=["POST"])
def submit_attendance():
    data = request.json
    emp_id = str(data.get("emp_id", "")).strip()
    action = data.get("action")
    date = data.get("date")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Load master data and normalize employee ID column
    df_master = pd.read_excel(MASTER_DATA_PATH)
    df_master["Employee ID"] = df_master["Employee ID"].astype(str).str.strip()
    employee = df_master[df_master["Employee ID"] == emp_id]

    if employee.empty:
        return jsonify({"remark": "Employee ID not found."})

    employee_data = employee.iloc[0]
    emp_name = employee_data["Full Name"]
    department = employee_data["Department"]
    position = employee_data["Position"]
    status = employee_data["Status"]

    # Load or create Attendance_Log.xlsx
    if os.path.exists(ATTENDANCE_LOG_PATH):
        book = load_workbook(ATTENDANCE_LOG_PATH)
    else:
        # Create new workbook with no sheets
        book = Workbook()
        # Remove default sheet if exists
        default_sheet = book.active
        book.remove(default_sheet)

    # If sheet exists for the date, load it
    if date in book.sheetnames:
        sheet = book[date]
        data_rows = list(sheet.values)
        if data_rows:
            df_log = pd.DataFrame(data_rows)
            df_log.columns = df_log.iloc[0]
            df_log = df_log[1:]
        else:
            # Empty sheet, create empty dataframe with headers
            headers = ["Employee ID", "Full Name", "Department", "Position", "Status", "IN time", "OUT time"]
            df_log = pd.DataFrame(columns=headers)
    else:
        # Create new sheet with headers
        sheet = book.create_sheet(date)
        headers = ["Employee ID", "Full Name", "Department", "Position", "Status", "IN time", "OUT time"]
        sheet.append(headers)
        df_log = pd.DataFrame(columns=headers)

    # Normalize Employee ID in log for comparison
    if not df_log.empty and "Employee ID" in df_log.columns:
        df_log["Employee ID"] = df_log["Employee ID"].astype(str).str.strip()
        existing = df_log[df_log["Employee ID"] == emp_id]
    else:
        existing = pd.DataFrame()

    remark = ""

    if action == "IN":
        if not existing.empty and pd.notna(existing.iloc[0]["IN time"]) and existing.iloc[0]["IN time"] != "":
            remark = "IN already marked."
        else:
            new_row = [emp_id, emp_name, department, position, status, timestamp, ""]
            sheet.append(new_row)
            remark = "IN marked."

    elif action == "OUT":
        if existing.empty:
            remark = "No IN found. Cannot mark OUT."
        elif pd.notna(existing.iloc[0]["OUT time"]) and existing.iloc[0]["OUT time"] != "":
            remark = "OUT already marked."
        else:
            # Update OUT time in the sheet for the employee
            for row in sheet.iter_rows(min_row=2):
                if str(row[0].value).strip() == emp_id:
                    row[6].value = timestamp  # OUT time is 7th column (index 6)
                    break
            remark = "OUT marked."
    else:
        remark = "Invalid action."

    book.save(ATTENDANCE_LOG_PATH)
    return jsonify({"remark": remark})

if __name__ == "__main__":
    app.run(debug=True)
